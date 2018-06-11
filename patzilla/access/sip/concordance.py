# -*- coding: utf-8 -*-
# (c) 2014-2018 Andreas Motl <andreas.motl@ip-tools.org>
import sys
import logging

from tqdm import tqdm
from csv import DictReader
from openpyxl.reader.excel import load_workbook
from openpyxl.exceptions import InvalidFileException
from pymongo.uri_parser import parse_uri
from mongoengine import connect as mongoengine_connect
from mongoengine.document import Document
from mongoengine.fields import StringField, IntField
from patzilla.util.ipc.parser import IpcDecoder

log = logging.getLogger(__name__)


# ------------------------------------------
#   bootstrapping
# ------------------------------------------
def includeme(config):
    config.add_subscriber(setup_mongoengine, "pyramid.events.ApplicationCreated")


# ------------------------------------------
#   database connection
# ------------------------------------------
def setup_mongoengine(event):
    registry = event.app.registry
    mongodb_uri = registry.settings.get('mongodb.patzilla.uri')
    mongodb_database = parse_uri(mongodb_uri)['database']
    mongoengine_connect(mongodb_database, host=mongodb_uri)


# ------------------------------------------
#   data model
# ------------------------------------------
class SipCountry(Document):
    ccid = IntField(unique=True)
    cc = StringField(unique=True)

    meta = {
        'indexes': ['ccid', 'cc']
    }

class SipIpcClass(Document):
    itid = IntField(unique=True)
    ipc = StringField()

    meta = {
        'indexes': ['ipc']
    }

class SipCpcClass(Document):
    cpcid = IntField(unique=True)
    cpc = StringField()

    meta = {
        'indexes': ['cpc']
    }


# ------------------------------------------
#   import functions
# ------------------------------------------
def import_countries(xlsfile, force=False):

    """
    countries = [
        SipCountry(ccid=1, cc='AP'),
        SipCountry(ccid=2, cc='AR'),
        ]
    for country in countries:
        country.save()
    """

    log.info('SIP country map: Starting import')

    count = SipCountry.objects().count()
    if count > 0:
        if force:
            log.info('SIP country map: Dropping collection because of "--force"')
            SipCountry.drop_collection()
        else:
            log.info('SIP country map: Database contains {0} entries, will not import again'.format(count))
            return

    log.info('SIP country map: Opening concordance file {}'.format(xlsfile))
    try:
        wb = load_workbook(filename=xlsfile)
    except InvalidFileException as ex:
        log.error('SIP country map: Import failed, exception is ' + ex.message)
        return

    ws = wb.active
    for row in tqdm(ws.rows[1:]):
        #print "row:", row
        ccid = row[0].value
        cc = row[1].value
        #print ccid, cc
        country = SipCountry(ccid=ccid, cc=cc)
        country.save()

    count = SipCountry.objects().count()
    log.info('SIP country map: Imported {0} entries'.format(count))


def import_ipc_classes(xlsfile, force=False):

    log.info('SIP IPC class map: Starting import, this might take some time')

    count = SipIpcClass.objects().count()
    if count > 0:
        if force:
            log.info('SIP IPC class map: Dropping collection because of "--force"')
            SipIpcClass.drop_collection()
        else:
            log.info('SIP IPC class map: Database contains {0} entries, will not import again'.format(count))
            return

    terminator_chars = ['0', '-']

    log.info('SIP IPC class map: Opening concordance file {}'.format(xlsfile))
    try:
        wb = load_workbook(filename=xlsfile)
    except InvalidFileException as ex:
        log.error('SIP IPC class map: Reading XLSX file {} failed: {}'.format(xlsfile, ex.message))
        return

    log.info('SIP IPC class map: Importing data')
    ws = wb.active
    for row in tqdm(ws.rows[1:]):
        #print "row:", row
        itid = row[0].value
        c1 = str(row[1].value)
        c2 = str(row[2].value)
        c3 = str(row[3].value)
        c4 = str(row[4].value)
        c5 = str(row[5].value)

        ipc_dict = IpcDecoder.getempty()
        ipc_dict['section'] = c1
        if c2 not in terminator_chars:
            ipc_dict['class'] = c2
        if c3 not in terminator_chars:
            ipc_dict['subclass'] = c3
        if c4 not in terminator_chars:
            ipc_dict['group'] = c4
        if c5 not in terminator_chars:
            ipc_dict['subgroup'] = c5

        #print ipc_dict
        ipc = IpcDecoder(ipc_dict=ipc_dict)
        ipc_docdb = ipc.formatOPS()
        #print 'ipc-docdb:', ipc_docdb

        ipc_class = SipIpcClass(itid=itid, ipc=ipc_docdb)
        ipc_class.save()

    count = SipIpcClass.objects().count()
    log.info('SIP IPC class map: Imported {0} entries'.format(count))


def import_cpc_classes(filename, force=False):
    """
    mdb-schema IPC_CPC.mdb
    mdb-export IPC_CPC.mdb cpcterm > IPC_CPC.csv
    """

    log.info('SIP CPC class map: Starting import, this might take some time')

    count = SipCpcClass.objects().count()
    if count > 0:
        if force:
            log.info('SIP CPC class map: Dropping collection because of "--force"')
            SipCpcClass.drop_collection()
        else:
            log.info('SIP CPC class map: Database contains {0} entries, will not import again'.format(count))
            return

    terminator_chars = ['0', '-']

    log.info('SIP CPC class map: Opening concordance file {}'.format(filename))

    if filename.endswith('.csv'):

        def decode_row(row):
            cpcid = row['ID']
            c1 = str(row['C1'])
            c2 = str(row['C2'])
            c3 = str(row['C3'])
            c4 = str(row['C4'])
            c5 = str(row['C5'])
            return locals()

        try:
            csvfile = open(filename)
        except IOError as ex:
            log.error('SIP CPC class map: Opening file {} failed: {}'.format(filename, str(ex)))
            return

        try:
            stream = DictReader(csvfile)
            print stream.fieldnames
        except Exception as ex:
            log.error('SIP CPC class map: Reading CSV file {} failed: {}'.format(filename, ex.message))
            return

    elif filename.endswith('.xlsx'):

        def decode_row(row):
            cpcid = row[0].value
            c1 = str(row[1].value)
            c2 = str(row[2].value)
            c3 = str(row[3].value)
            c4 = str(row[4].value)
            c5 = str(row[5].value)
            return locals()

        try:
            wb = load_workbook(filename=filename)
        except InvalidFileException as ex:
            log.error('SIP IPC class map: Reading XLSX file {} failed: {}'.format(filename, ex.message))
            return

        ws = wb.active
        print 'XLSX row 1:', [cell.value for cell in ws.rows[0]]
        stream = ws.rows[1:20]

    #sys.exit(1)

    log.info('SIP CPC class map: Importing data')
    for row in tqdm(stream, total=255628):
        item = decode_row(row)
        #print "row:", row
        #print "item:", item
        cpcid = int(item['cpcid'])
        c1 = item['c1']
        c2 = item['c2']
        c3 = item['c3']
        c4 = item['c4']
        c5 = item['c5']

        ipc_dict = IpcDecoder.getempty()
        ipc_dict['section'] = c1
        if c2 not in terminator_chars:
            ipc_dict['class'] = c2
        if c3 not in terminator_chars:
            ipc_dict['subclass'] = c3
        if c4 not in terminator_chars:
            ipc_dict['group'] = c4
        if c5 not in terminator_chars:
            ipc_dict['subgroup'] = c5

        #print ipc_dict
        ipc = IpcDecoder(ipc_dict=ipc_dict)
        ipc_docdb = ipc.formatOPS()
        #print 'ipc-docdb:', ipc_docdb

        ipc_class = SipCpcClass(cpcid=cpcid, cpc=ipc_docdb)
        ipc_class.save()

    count = SipCpcClass.objects().count()
    log.info('SIP CPC class map: Imported {0} entries'.format(count))
